"""
Convert Google Sheet 'du an chay' (200 store) -> import_batch.json.
Ap dung template Coupon (tab 'cach len tk') voi placeholder {Store} va {%}.

Usage:
    python convert_sheet_to_import.py
        --sheet "c:/tmp/user_sheet.xlsx"
        --out   "c:/Users/Admin/Documents/vippro campping/batch_200.json"
        --default-rate 55
"""
from __future__ import annotations
import argparse, json, re, sys, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


# ---------- Templates (rut tu tab 'cach len tk') ----------
# Placeholder: {Store} va {%}  (raw text goc dung "Store" va "%")

KEYWORDS_COUPON = [
    '"Store coupon"', '"Store coupon code"', '"Store promo code"', '"Store promo"',
    '"Promo code Store"', '"Coupon Store"', '"Coupon code Store"',
    '"Coupon code for Store"', '"Coupon for Store"', '"Promo code for Store"',
]

HEADLINES_COUPON = [
    "Store Coupon Code", "Best Store Coupons", "Top Store Coupons", "Get % Off Store",
    "Top Coupon Codes Today", "Enjoy Fast Savings % Off", "All Codes [Verified]",
    "Save % Off Coupon Code",
]

DESCRIPTIONS_COUPON = [
    "Save % Off on all products with Store coupon. All coupons verified",
    "With This Exclusive Store Coupon Code, Shop Till You Drop And Enjoy % Off",
    "Saving Your Pocket With % Off Thanks To Special Store Coupon. Enjoy It",
    "Enjoy A Big Surprise With % Off Thanks To lam Store. Limited Offer!",
    "Get Store Promo Code At Checkout And Enjoy % Off. Don't Hesitate Anymore.",
    "% Off Store Coupons & Promo Codes 2025",
    "Apply these Store Coupon At Checkout And Save. Tested Daily.",
]

# 3 sitelink co dinh cho moi camp
SITELINK_TEMPLATES = [
    {"text": "Today's Top Coupon",    "url_suffix": "&1", "desc1": "Provide the best coupons", "desc2": "Rated by the user"},
    {"text": "Best Deals Of The Days", "url_suffix": "",   "desc1": "Sale Up To 60%",          "desc2": "To save your money"},
    {"text": "All Blog",               "url_suffix": "&2", "desc1": "Save with All Blog coupons", "desc2": "Discounts and promo codes for"},
]

CALLOUTS_DEFAULT = ["Free Shipping", "100% Working Code", "Verified Daily", "Top Deals"]

BUDGET_DEFAULT = "10"
BIDDING_DEFAULT = "Tối đa lượt nhấn chuột"
GENDER_DEFAULT = "Tất cả"
AGE_DEFAULT = ["18-24", "25-34", "35-44", "45-54", "55-64", "65+"]
DEVICES_DEFAULT = ["Di động", "Máy tính", "Máy tính bảng"]


# ---------- Helpers ----------
def apply_template(tpl: str, store: str, rate_pct: str) -> str:
    """Replace 'Store' -> ten store, '%' -> '55%' (giu dau %)."""
    return tpl.replace("Store", store).replace("%", f"{rate_pct}%")


COUNTRY_RE = re.compile(r"^\s*([A-Z]{2})\b")

def parse_country(cell: str | None) -> str | None:
    """Tu 'US (13,744)' -> 'US'. Tra ve None neu khong match."""
    if not cell:
        return None
    m = COUNTRY_RE.match(str(cell))
    return m.group(1) if m else None


# Map country code -> ten VN dung trong GG Ads UI (Selenium type vao search box)
COUNTRY_VN = {
    "US": "Hoa Kỳ", "UK": "Vương quốc Anh", "GB": "Vương quốc Anh",
    "CA": "Canada", "AU": "Úc", "DE": "Đức", "FR": "Pháp",
    "IT": "Ý", "ES": "Tây Ban Nha", "NL": "Hà Lan", "PL": "Ba Lan",
    "JP": "Nhật Bản", "KR": "Hàn Quốc", "PH": "Philippines",
    "ID": "Indonesia", "MY": "Malaysia", "TH": "Thái Lan",
    "SG": "Singapore", "VN": "Việt Nam", "BR": "Brazil",
    "MX": "Mexico", "AR": "Argentina", "TR": "Thổ Nhĩ Kỳ",
    "SA": "Ả Rập Xê Út", "AE": "Các Tiểu vương quốc Ả Rập Thống nhất",
    "ZA": "Nam Phi", "EG": "Ai Cập", "RU": "Nga", "UA": "Ukraine",
    "SE": "Thụy Điển", "NO": "Na Uy", "FI": "Phần Lan", "DK": "Đan Mạch",
    "BE": "Bỉ", "AT": "Áo", "CH": "Thụy Sĩ", "IE": "Ireland",
    "PT": "Bồ Đào Nha", "CZ": "Cộng hòa Séc", "HU": "Hungary",
    "RO": "Romania", "GR": "Hy Lạp", "NZ": "New Zealand",
    "HK": "Hồng Kông", "TW": "Đài Loan", "CL": "Chile", "CO": "Colombia",
    "PE": "Peru", "IL": "Israel", "CN": "Trung Quốc",
    # IN bi bo (theo ghi chu sheet)
}

def build_target_locations(top1: str | None, top2: str | None, top3: str | None) -> list[str]:
    """Top 3 country, strip '(...)', skip IN, map sang ten VN."""
    codes = []
    for cell in (top1, top2, top3):
        code = parse_country(cell)
        if code and code != "IN" and code not in codes:
            codes.append(code)
    return [COUNTRY_VN.get(c, c) for c in codes]


def build_camp_config(store: str, link: str, rate_pct: str,
                      top1: str, top2: str, top3: str) -> dict:
    """Sinh 1 camp config cho 1 store."""
    # Clean rate: '55%' hoac '55' -> '55'
    rate_clean = str(rate_pct).replace("%", "").strip() or "55"

    headlines = [apply_template(h, store, rate_clean) for h in HEADLINES_COUPON]
    descriptions = [apply_template(d, store, rate_clean) for d in DESCRIPTIONS_COUPON]
    keywords = [apply_template(k, store, rate_clean) for k in KEYWORDS_COUPON]

    # Sitelinks: url = link + suffix
    sitelinks = []
    for sl in SITELINK_TEMPLATES:
        sitelinks.append({
            "text": sl["text"],
            "url": f"{link}{sl['url_suffix']}",
            "desc1": sl["desc1"],
            "desc2": sl["desc2"],
        })

    target = build_target_locations(top1, top2, top3)

    return {
        "name": store,
        "campaignType": "Search",
        "link1": link,
        "link2": "",
        "adsKey": keywords,
        "bidding": BIDDING_DEFAULT,
        "cpc": "",
        "budget": BUDGET_DEFAULT,
        "targetLocations": target,
        "excludeLocations": [],
        "devices": DEVICES_DEFAULT,
        "ageRange": AGE_DEFAULT,
        "gender": GENDER_DEFAULT,
        "headlines": headlines,
        "descriptions": descriptions,
        "sitelinks": sitelinks,
        "callouts": CALLOUTS_DEFAULT,
        "status": "pending",
        "notes": f"Auto-gen tu sheet 'du an chay' | rate={rate_clean}%",
    }


# ---------- Main ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", default=r"c:/tmp/user_sheet.xlsx")
    ap.add_argument("--out",   default=r"c:/Users/Admin/Documents/vippro campping/batch_200.json")
    ap.add_argument("--default-rate", default="55", help="Ty le mac dinh neu sheet chua co cot Ty Le")
    ap.add_argument("--rate-col", default=None, help="Ten cot chua ty le")
    ap.add_argument("--tab", default="dự án chạy")
    args = ap.parse_args()

    wb = load_workbook(args.sheet, data_only=True)
    if args.tab not in wb.sheetnames:
        print(f"ERROR: sheet khong co tab '{args.tab}'. Co: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[args.tab]

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    print(f"Header ({len(header)} cot): {header}")

    # Tim index cac cot bat buoc
    def idx(name):
        for i, h in enumerate(header):
            if h.lower() == name.lower():
                return i
        return -1

    i_store = idx("Cửa hàng")
    i_link = idx("Link Web Coupon")
    i_top1 = idx("Top 1 Country")
    i_top2 = idx("Top 2 Country")
    i_top3 = idx("Top 3 Country")
    i_rate = idx(args.rate_col) if args.rate_col else idx("Tỷ Lệ")
    if i_rate < 0:
        i_rate = idx("Ty Le")

    if i_store < 0 or i_link < 0:
        print(f"ERROR: thieu cot 'Cửa hàng' hoac 'Link Web Coupon'")
        sys.exit(1)

    camps = []
    skipped = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or not row[i_store]:
            continue
        store = str(row[i_store]).strip()
        link = str(row[i_link] or "").strip()
        if not store or not link:
            skipped.append((row_num, store, "thieu ten hoac link"))
            continue

        rate = str(row[i_rate]).strip() if i_rate >= 0 and row[i_rate] else args.default_rate
        top1 = str(row[i_top1] or "") if i_top1 >= 0 else ""
        top2 = str(row[i_top2] or "") if i_top2 >= 0 else ""
        top3 = str(row[i_top3] or "") if i_top3 >= 0 else ""

        camp = build_camp_config(store, link, rate, top1, top2, top3)
        camps.append(camp)

    # Xuat JSON
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(camps, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Da sinh {len(camps)} camp -> {out_path}")
    print(f"  Size: {out_path.stat().st_size:,} bytes")
    if skipped:
        print(f"  Skip {len(skipped)} dong:")
        for r, s, why in skipped[:10]:
            print(f"    row {r} ({s}): {why}")

    # Preview 1 camp dau
    if camps:
        print(f"\n=== PREVIEW camp 1 ({camps[0]['name']}) ===")
        preview = {k: v for k, v in camps[0].items()
                   if k in ("name", "link1", "targetLocations", "headlines", "sitelinks", "notes")}
        print(json.dumps(preview, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
